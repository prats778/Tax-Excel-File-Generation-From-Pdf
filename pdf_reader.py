import pandas as pd
import re
import openpyxl
from openpyxl.styles import PatternFill
from PyPDF2 import PdfReader

class pdfreader():
    def __init__(self,path):
        self.year = None
        self.gst_num = None
        self.periods = []
        self.tax_values = []
        self.path = path
        self.text_extractor()

    def text_extractor(self):
        with open(self.path, 'rb') as f:
            pdf = PdfReader(f)
            number_of_pages = len(pdf.pages)
            for i in range(0,number_of_pages,2):
                page_num = i+1
                print("Processing page ",page_num," ====> ")     
                # get the ith page
                page = pdf.pages[i]
                text = page.extractText()
                # print("Text: ==> ",text)
                self.process_text_iteratively(text)
                print(" <======= Finished processing ",page_num)
        
        self.export_data()
            
    def process_text_iteratively(self,text):
        lines = text.split('\n')  # Split the text into lines
        found_tax = False
        #Extract Period
        period_match = re.search(r'Period (\w+)', text)
        period = period_match.group(1) if period_match else None
        self.periods.append(period)

        for line in lines:
            # Extract year
            if not self.year and 'Year' in line:
                year_match = re.search(r'(\d{4}-\d{2})', line)
                if year_match:
                    year = year_match.group(1)
                    self.year = year
            
            # Extract GST number
            elif not self.gst_num and 'GSTIN of the supplier' in line:
                gst_number_match = re.search(r'(\w+)$', line)
                if gst_number_match:
                    gst_number = gst_number_match.group(1)
                    self.gst_num = gst_number

            # Extract taxable value after specific keywords/phrases
            elif not found_tax and 'exempted)' in line:
                taxable_value_match = re.search(r'([\d.,]+)', line)
                if taxable_value_match:
                    total_taxable_value = taxable_value_match.group(1)
                    self.tax_values.append(total_taxable_value)
                    found_tax = True  

    def export_data(self):

        # Preparing data for DataFrame
        data = {
            "Month": self.periods,
            "Taxable Value": self.tax_values
        }

        print("data ==> ", data)
        # Creating DataFrame
        df = pd.DataFrame(data)

        # Convert 'total_taxable_value' to numeric, handling non-numeric with 'coerce'
        df['Taxable Value'] = pd.to_numeric(df['Taxable Value'], errors='coerce')

        # Optional: Fill NaN values if any
        df['Taxable Value'] = df['Taxable Value'].fillna(0)

        # Saving DataFrame to Excel without header and index
        excel_path = f"{self.gst_num}.xlsx"
        df.to_excel(excel_path, index=False, header=False, engine='openpyxl', startrow=4)

        # Opening the Excel file and adding GST Number, Year, and headers manually
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active

        # Define fills
        grayFill = PatternFill(start_color='00CCCCCC',
                                end_color='00CCCCCC',
                                fill_type='solid')
        blueFill = PatternFill(start_color='0099CCFF',
                            end_color='0099CCFF',
                            fill_type='solid')

        # Insert GST Number and Year with background color
        sheet['A1'] = f"GST Number:"
        sheet['B1'] = f"{self.gst_num}"
        sheet['A1'].fill = grayFill
        sheet['A2'] = f"Year:"
        sheet['B2'] = f"{self.year}"
        sheet['A2'].fill = grayFill

        # Adding column headers with background color
        sheet['A4'] = "Month"
        sheet['A4'].fill = blueFill
        sheet['B4'] = "Taxable Value"
        sheet['B4'].fill = blueFill


        sheet.column_dimensions['A'].width = 15.
        sheet.column_dimensions['B'].width = 15.

        # Save the changes to the Excel file
        workbook.save(excel_path)
        workbook.close()

        print(" ===================== Data saved to excel file successfully. ===================")

    
